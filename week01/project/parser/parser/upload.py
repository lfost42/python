"""
Uploads excel file and parses date from file name.
"""
import os
import os.path
import datetime
import re
from openpyxl import load_workbook
from handler import app
# import config as cfg

# cfg = app.config.from_pyfile('config.py', instance_relative_config=True)
# with app.open_instance_resource('application.cfg') as f:
#     config = f.read()

def check_file(file):
    """Check if file is a duplicate.

    Args:
        file (FileStorage): file for upload
    """
    path = f'files/archived/{file.filename}'

    if os.path.exists(path):
        app.logger.error(f'{file} duplicate file | check_file')
        return -1
    app.logger.info('{file} checked, not a duplicate.')
    return 1

def read_file(file):
    """Reads file

    Args:
        file (FileStorage): file for upload
    """
    try:
        workbook = load_workbook(file, data_only=True)
        app.logger.info(f'{file} read.')
        return workbook
    except ImportError as import_error:
        app.logger.exception(f'Data uanble to load: {import_error} | read_file')
        return -1

def archive_file(file):
    """Move to archived folder

    Args:
        file (FileStorage): file for upload
    """
    arch_folder = 'files/archived/'

    if os.path.exists(arch_folder):
        app.logger.info(f'{file} to archived folder.')
        file.save(os.path.join(arch_folder, file.filename))

def error_file(file):
    """Move to error folder

    Args:
        file (FileStorage): file for upload
    """
    err_folder = 'files/error/'

    if os.path.exists(err_folder):
        app.logger.info(f'{file} moved to error folder.')
        file.save(os.path.join(err_folder, file.filename))

def find_row(worksheet, search):
    """Uses search term to find a column coordinate.

    Args:
        worksheet (FileStorage): The active worksheet in the excel file we are parsing.
        search (string): Search term required to lcoate a row.

    Returns:
        Integer: Coordinate for searched row.
    """
    for my_row in range(1, worksheet.max_row + 1):
        text = worksheet.cell(row = my_row, column = 1).value

        if text is not None:
            if isinstance(text, datetime.datetime):
                date_str = str(text.strftime("%b-%Y")).lower()
                new_txt = date_str[:4] + date_str[-2:]
                if new_txt == search:
                    app.logger.info(f'{search} found in {worksheet}.')
                    return my_row

            if isinstance(text, str):
                if search.lower() in text.lower():
                    app.logger.info(f'{search} found in {worksheet}.')
                    return my_row

    app.logger.error(f'{search} not found in {worksheet} | find_row')
    return -1

def find_column(worksheet, search):
    """Uses search term to find a column coordinate.

    Args:
        worksheet (FileStorage): The active worksheet in the excel file we are parsing.
        search (string): Search term required to locate a column.

    Returns:
        Integer: Coordinate for searched column.
    """
    for my_col in range(1, worksheet.max_column + 1):
        text = worksheet.cell(row = 1, column = my_col).value

        if text is not None:
            if not isinstance(text, datetime.datetime):
                if isinstance(text, str):
                    if search.lower() in text.lstrip().lower():
                        app.logger.info(f'{search} located in {worksheet}.')
                        return my_col

            if isinstance(text, datetime.datetime):
                date_str = str(text.strftime("%b"))
                if search.lower() in date_str.lower():
                    app.logger.info(f'{date_str} located in {worksheet}')
                    return my_col

    app.logger.error(f'{search} not found in {worksheet} | find_column')
    return -1

def check_year(file):
    """Parses file name for a 4 digit number.

    Args:
        file (string): Name of excel file.

    Returns:
        string: The last 2 digits in the located 4 digit number.
    """
    filename = file.filename
    year = re.findall('[0-9]+', filename)

    if (not year or len(year[0]) != 4):
        error_file(file)
        app.logger.error(f'Could not extract {year} | check_year')
        return -1

    app.logger.info('Year successfully parsed form file name.')
    return year[0][-2:]

def check_month(file):
    """Parses file name for month.

    Args:
        file (string): Name of the file.

    Returns:
        string: First 3 characters to the month found.
        integer: -1 if month could not be extracted.
    """
    filename = file.filename.lower()
    months = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']

    for month in months:
        if month in filename:
            app.logger.info(f'Month successfully parsed from {file}.')
            return month
    app.logger.error(f'ERROR: unable to extract month from {file}| check_month')
    return -1

def get_date(file):
    """Gets date from file.

    Args:
        file (string): Name of file.

    Returns:
        list: List with year at index 0 and month at index 1.
    """
    year = check_year(file)
    month = check_month(file)

    if year == -1 or month == -1:
        app.logger.info('This should not happen.')
        return -1

    app.logger.info(f'Date generated from {file}')
    return [year, month]
