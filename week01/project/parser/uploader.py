import os
from openpyxl import load_workbook
from openpyxl.utils import rows_from_range
from flask import Flask
import unicodedata
import datetime

app = Flask(__name__)

def check_file(file):
    """check if file matches expedia_report_summary

    Args:
        file (FileStorage): the file we are attempting to upload
    """
    if not file.filename.startswith('expedia_report_monthly_'):
        error_file(file)
        app.logger.error(f'{file} moved to error, file name does not start with expedia_report_monthly_')

def read_file(file):
    """_summary_

    Args:
        file (_type_): _description_

    Returns:
        _type_: _description_
    """
    try:
        if not file:
            print('No file found with name: expedia_report_monthly_MONTH_YEAR.xlsx')
        else:
            try:
                workbook = load_workbook(file, data_only=True)
                return workbook
            except:
                app.logger.error('Data uanble to load')

    except:
        app.logger.error('Error reading File')

def archive_file(file):
    """Move to archived folder

    Args:
        file (FileStorage): the file we are attempting to upload
    """
    arch_folder = 'files/archived'
    if os.path.exists(arch_folder):
        file.save(os.path.join(arch_folder, file.filename))
    return 'archived'

def error_file(file):
    """Move to error folder

    Args:
        file (_type_): FileStorage
    """
    err_folder = 'files/error/'
    if os.path.exists(err_folder):
        file.save(os.path.join(err_folder, file.filename))
    return 'error'

def find_row(worksheet, search):
    for my_row in range(1, worksheet.max_row + 1):
        text = worksheet.cell(row = my_row, column = 1).value
        
        if text is not None:
            if type(text) == datetime.datetime:
                date_str = str(text.strftime("%b-%Y")).lower()
                new_txt = date_str[:4]+date_str[-2:]
                
                if new_txt == search:
                    return my_row
                
            if isinstance(text, str):
                if search.lower() in text.lower():
                    return my_row
                
    app.logger.error(f'{search} not found (row)')
    return -1

def find_column(worksheet, search):    
    for my_col in range(1, worksheet.max_column + 1):
        text = worksheet.cell(row = 1, column = my_col).value
            
        if text is not None:
            if type(text) != datetime.datetime:
                if isinstance(text, str):
                    if search.lower() in text.lstrip().lower():
                        return my_col
            
            if type(text) == datetime.datetime:
                date_str = str(text.strftime("%b"))
                if search.lower() in date_str.lower():
                    app.logger.critical(f'returning my column: {my_col}')
                    return my_col

    app.logger.error(f'{search} not found (column)')
    return -1